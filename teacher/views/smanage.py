from django.db.models import Count, Q
from utils.encrypt import md5
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from teacher import models
from utils.pagination import Pagination
from openpyxl import load_workbook
from teacher.form import LoginForm, RegisterModelForm, ArticleForm
from student.models import StudentInfo, StudentArticles
from student.form import ArticleForm

def student_list(request):
    value = request.GET.get('q', '')  # 如果不能获取到q的值就是空字符串

    queryset = StudentInfo.objects.filter(
        Q(identitycard__contains=value) | Q(name__contains=value) | Q(class_room__contains=value))

    if queryset:
        page_obj = Pagination(request, queryset, page_size=20, plus=3)
        page_queryset = page_obj.page_queryset
        page_string = page_obj.html()
        prama_dict = {
            'value': value,
            'page_queryset': page_queryset,
            'page_string': page_string,
            'count': queryset.count(),
        }
        return render(request, "teacher/student_list.html", prama_dict)
    return render(request, "teacher/student_list.html")

@csrf_exempt
def student_batch(request):
    file_obj = request.FILES.get('file_obj', '')
    if file_obj == '':
        return JsonResponse({'status': False})

    wb = load_workbook(file_obj)

    sheet = wb.worksheets[0]
    row_dict = {}
    for row in sheet.iter_rows(min_row=2):
        row_dict["class_room"] = row[0].value
        row_dict["name"] = row[1].value
        if row[2].value == '男':
            row_dict["gender"] = 1
        else:
            row_dict["gender"] = 2
        row_dict["identitycard"] = row[3].value
        row_dict["password"] = md5(row[3].value[-6:])
        exists = StudentInfo.objects.filter(identitycard=row[3].value).exists()
        if not exists:
            StudentInfo.objects.create(**row_dict)
    return JsonResponse({'status': True})

def student_clear(request):
    StudentInfo.objects.all().delete()
    return redirect('/teacher/s_list')